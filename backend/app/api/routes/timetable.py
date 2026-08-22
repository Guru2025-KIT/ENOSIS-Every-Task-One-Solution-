import uuid
import io
import csv
from collections import defaultdict
from datetime import datetime, timezone
from typing import Any
import openpyxl

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session

from app.api.deps import get_current_user, require_timetable_manager
from app.core.config import settings
from app.db.base import get_db
from app.models.academic import (
    Division, Subject, Room, TeachingAssignment, FacultyUnavailability,
    InstitutionalCourse, SharedCourse
)
from app.models.timetable import TimetableEntry
from app.models.user import User
from app.models.schedule_config import ScheduleConfig
from app.models.constraints import TimetableConstraint
from app.models.generation_history import GenerationRun
from app.schemas.timetable import (
    DivisionCreate, DivisionOut,
    SubjectCreate, SubjectOut,
    RoomCreate, RoomOut,
    TeachingAssignmentCreate, TeachingAssignmentOut,
    FacultyUnavailabilityCreate,
    ScheduleConfigCreate, ScheduleConfigOut,
    ConstraintCreate, ConstraintOut,
    TimetableGenerationRequest, SolverDivision, SolverSubject, SolverRoom,
    SolverAssignment, SolverUnavailability, SolverSoftConstraint,
    TimetableEntryOut, CollegeInfo, GenerationRunOut, ValidationResponse,
    ConflictDetail,
    InstitutionalCourseCreate, InstitutionalCourseOut,
    SharedCourseCreate, SharedCourseOut
)
from app.services.timetable_solver import generate_timetable, validate_request
from app.services.timetable_validator import validate_generated_timetable
from app.services.notifications import notify

router = APIRouter(prefix="/timetable", tags=["timetable"])


@router.get("/college-info", response_model=CollegeInfo)
def college_info(db: Session = Depends(get_db)):
    """
    Public (no auth) — the app's splash/login/timetable screens display
    the institution's name. First checks if a database schedule configuration
    exists and has a college name, otherwise falls back to .env settings.
    """
    config = db.query(ScheduleConfig).filter(ScheduleConfig.id == "default").first()
    if config and config.college_name:
        return CollegeInfo(college_name=config.college_name)
    return CollegeInfo(college_name=settings.COLLEGE_NAME)


@router.get("/faculty-list")
def faculty_list(db: Session = Depends(get_db), _: User = Depends(require_timetable_manager)):
    """
    A minimal faculty picker for the assignment-creation screen.
    """
    faculty = db.query(User).all()
    return [{"id": u.id, "full_name": u.full_name, "email": u.email} for u in faculty]


# ---------------------------------------------------------------------------
# Global Schedule Configuration (CRUD)
# ---------------------------------------------------------------------------

@router.get("/schedule-config", response_model=ScheduleConfigOut)
def get_schedule_config(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    """Retrieves the active timetable schedule configuration (always uses id='default')."""
    config = db.query(ScheduleConfig).filter(ScheduleConfig.id == "default").first()
    if not config:
        # Auto-create a default configuration if none exists yet
        config = ScheduleConfig(
            id="default",
            working_days=6,
            day_names=["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"],
            periods_per_day=8,
            period_duration_minutes=60,
            lecture_duration_minutes=60,
            lab_duration_minutes=120,
            tutorial_duration_minutes=60,
            start_time="09:00",
            break_slots=[],
            break_labels={},
            max_lectures_per_day_per_faculty=None,
            college_name=settings.COLLEGE_NAME,
            department_name="Computer Science & Engineering",
            academic_year="2026-2027",
            semester="Odd",
            hod_name="Dr. Uma Gurav",
            time_limit_seconds=30
        )
        db.add(config)
        db.commit()
        db.refresh(config)
    return config


@router.post("/schedule-config", response_model=ScheduleConfigOut)
def update_schedule_config(
    payload: ScheduleConfigCreate,
    db: Session = Depends(get_db),
    _: User = Depends(require_timetable_manager)
):
    """Upserts the active schedule configuration."""
    config = db.query(ScheduleConfig).filter(ScheduleConfig.id == "default").first()
    if not config:
        config = ScheduleConfig(id="default")
        db.add(config)

    for field, value in payload.model_dump().items():
        setattr(config, field, value)

    db.commit()
    db.refresh(config)
    return config


# ---------------------------------------------------------------------------
# Generic Timetable Constraints (CRUD)
# ---------------------------------------------------------------------------

@router.post("/constraints", response_model=ConstraintOut, status_code=status.HTTP_201_CREATED)
def create_constraint(
    payload: ConstraintCreate,
    db: Session = Depends(get_db),
    _: User = Depends(require_timetable_manager)
):
    """Creates a new timetable constraint (either hard or soft)."""
    constraint = TimetableConstraint(**payload.model_dump())
    db.add(constraint)
    db.commit()
    db.refresh(constraint)
    return constraint


@router.get("/constraints", response_model=list[ConstraintOut])
def list_constraints(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    """Lists all active and inactive constraints saved in the database."""
    return db.query(TimetableConstraint).order_by(TimetableConstraint.created_at.desc()).all()


@router.delete("/constraints/{constraint_id}")
def delete_constraint(
    constraint_id: str,
    db: Session = Depends(get_db),
    _: User = Depends(require_timetable_manager)
):
    """Deletes a constraint by ID."""
    constraint = db.query(TimetableConstraint).filter(TimetableConstraint.id == constraint_id).first()
    if not constraint:
        raise HTTPException(status_code=404, detail="Constraint not found")
    db.delete(constraint)
    db.commit()
    return {"status": "deleted"}


# ---------------------------------------------------------------------------
# Setup Data Entities (CRUD)
# ---------------------------------------------------------------------------

@router.post("/divisions", response_model=DivisionOut, status_code=status.HTTP_201_CREATED)
def create_division(payload: DivisionCreate, db: Session = Depends(get_db), _: User = Depends(require_timetable_manager)):
    division = Division(**payload.model_dump())
    db.add(division)
    db.commit()
    db.refresh(division)
    return division


@router.get("/divisions", response_model=list[DivisionOut])
def list_divisions(
    year: int | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    query = db.query(Division)
    if year is not None:
        query = query.filter(Division.year == year)
    return query.order_by(Division.year, Division.division_code).all()


@router.put("/divisions/{division_id}", response_model=DivisionOut)
def update_division(
    division_id: str,
    payload: DivisionCreate,
    db: Session = Depends(get_db),
    _: User = Depends(require_timetable_manager)
):
    """Updates an existing division by ID."""
    division = db.query(Division).filter(Division.id == division_id).first()
    if not division:
        raise HTTPException(status_code=404, detail="Division not found")
    for field, value in payload.model_dump().items():
        setattr(division, field, value)
    db.commit()
    db.refresh(division)
    return division


@router.delete("/divisions/{division_id}")
def delete_division(
    division_id: str,
    db: Session = Depends(get_db),
    _: User = Depends(require_timetable_manager)
):
    """Deletes a division and its associated teaching assignments."""
    division = db.query(Division).filter(Division.id == division_id).first()
    if not division:
        raise HTTPException(status_code=404, detail="Division not found")
    # Cascade-delete assignments and timetable entries for this division
    db.query(TeachingAssignment).filter(TeachingAssignment.division_id == division_id).delete()
    db.query(TimetableEntry).filter(TimetableEntry.division_id == division_id).delete()
    db.delete(division)
    db.commit()
    return {"status": "deleted"}


@router.post("/subjects", response_model=SubjectOut, status_code=status.HTTP_201_CREATED)
def create_subject(payload: SubjectCreate, db: Session = Depends(get_db), _: User = Depends(require_timetable_manager)):
    subject = Subject(**payload.model_dump())
    db.add(subject)
    db.commit()
    db.refresh(subject)
    return subject


@router.get("/subjects", response_model=list[SubjectOut])
def list_subjects(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return db.query(Subject).all()


@router.put("/subjects/{subject_id}", response_model=SubjectOut)
def update_subject(
    subject_id: str,
    payload: SubjectCreate,
    db: Session = Depends(get_db),
    _: User = Depends(require_timetable_manager)
):
    """Updates an existing subject."""
    subject = db.query(Subject).filter(Subject.id == subject_id).first()
    if not subject:
        raise HTTPException(status_code=404, detail="Subject not found")
    for field, value in payload.model_dump().items():
        setattr(subject, field, value)
    db.commit()
    db.refresh(subject)
    return subject


@router.delete("/subjects/{subject_id}")
def delete_subject(
    subject_id: str,
    db: Session = Depends(get_db),
    _: User = Depends(require_timetable_manager)
):
    """Deletes a subject and its associated teaching assignments."""
    subject = db.query(Subject).filter(Subject.id == subject_id).first()
    if not subject:
        raise HTTPException(status_code=404, detail="Subject not found")
    db.query(TeachingAssignment).filter(TeachingAssignment.subject_id == subject_id).delete()
    db.query(TimetableEntry).filter(TimetableEntry.subject_id == subject_id).delete()
    db.delete(subject)
    db.commit()
    return {"status": "deleted"}


@router.post("/rooms", response_model=RoomOut, status_code=status.HTTP_201_CREATED)
def create_room(payload: RoomCreate, db: Session = Depends(get_db), _: User = Depends(require_timetable_manager)):
    room = Room(**payload.model_dump())
    db.add(room)
    db.commit()
    db.refresh(room)
    return room


@router.get("/rooms", response_model=list[RoomOut])
def list_rooms(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return db.query(Room).all()


@router.put("/rooms/{room_id}", response_model=RoomOut)
def update_room(
    room_id: str,
    payload: RoomCreate,
    db: Session = Depends(get_db),
    _: User = Depends(require_timetable_manager)
):
    """Updates an existing room."""
    room = db.query(Room).filter(Room.id == room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="Room not found")
    for field, value in payload.model_dump().items():
        setattr(room, field, value)
    db.commit()
    db.refresh(room)
    return room


@router.delete("/rooms/{room_id}")
def delete_room(
    room_id: str,
    db: Session = Depends(get_db),
    _: User = Depends(require_timetable_manager)
):
    """Deletes a room."""
    room = db.query(Room).filter(Room.id == room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="Room not found")
    db.delete(room)
    db.commit()
    return {"status": "deleted"}


@router.post("/assignments", response_model=TeachingAssignmentOut, status_code=status.HTTP_201_CREATED)
def create_assignment(payload: TeachingAssignmentCreate, db: Session = Depends(get_db), _: User = Depends(require_timetable_manager)):
    for model, field_id, label in [
        (User, payload.faculty_id, "faculty"),
        (Subject, payload.subject_id, "subject"),
        (Division, payload.division_id, "division"),
    ]:
        if db.query(model).filter(model.id == field_id).first() is None:
            raise HTTPException(status_code=404, detail=f"No {label} found with id {field_id}")

    assignment = TeachingAssignment(**payload.model_dump())
    db.add(assignment)
    db.commit()
    db.refresh(assignment)
    return assignment


@router.get("/assignments", response_model=list[TeachingAssignmentOut])
def list_assignments(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return db.query(TeachingAssignment).all()


@router.get("/assignments-detailed")
def list_assignments_detailed(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    """Returns assignments with resolved faculty name, subject name, and division name for display."""
    assignments = db.query(TeachingAssignment).all()
    result = []
    for a in assignments:
        result.append({
            "id": a.id,
            "faculty_id": a.faculty_id,
            "faculty_name": a.faculty.full_name if a.faculty else "Unknown",
            "subject_id": a.subject_id,
            "subject_name": a.subject.name if a.subject else "Unknown",
            "division_id": a.division_id,
            "division_name": f"Year {a.division.year} - Div {a.division.division_code}" if a.division else "Unknown",
        })
    return result


@router.delete("/assignments/{assignment_id}")
def delete_assignment(
    assignment_id: str,
    db: Session = Depends(get_db),
    _: User = Depends(require_timetable_manager)
):
    """Deletes a teaching assignment by ID."""
    assignment = db.query(TeachingAssignment).filter(TeachingAssignment.id == assignment_id).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    db.delete(assignment)
    db.commit()
    return {"status": "deleted"}


@router.post("/unavailability", status_code=status.HTTP_201_CREATED)
def create_unavailability(payload: FacultyUnavailabilityCreate, db: Session = Depends(get_db), _: User = Depends(require_timetable_manager)):
    entry = FacultyUnavailability(**payload.model_dump())
    db.add(entry)
    db.commit()
    return {"status": "created"}


# ---------------------------------------------------------------------------
# Institutional Course CRUD
# ---------------------------------------------------------------------------

@router.post("/institutional-courses", response_model=InstitutionalCourseOut, status_code=status.HTTP_201_CREATED)
def create_institutional_course(
    payload: InstitutionalCourseCreate,
    db: Session = Depends(get_db),
    _: User = Depends(require_timetable_manager)
):
    if payload.faculty_id and db.query(User).filter(User.id == payload.faculty_id).first() is None:
        raise HTTPException(status_code=404, detail=f"No faculty found with id {payload.faculty_id}")
    if payload.room_id and db.query(Room).filter(Room.id == payload.room_id).first() is None:
        raise HTTPException(status_code=404, detail=f"No room found with id {payload.room_id}")

    course = InstitutionalCourse(**payload.model_dump())
    db.add(course)
    db.commit()
    db.refresh(course)
    return course


@router.get("/institutional-courses", response_model=list[InstitutionalCourseOut])
def list_institutional_courses(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return db.query(InstitutionalCourse).all()


@router.delete("/institutional-courses/{course_id}")
def delete_institutional_course(
    course_id: str,
    db: Session = Depends(get_db),
    _: User = Depends(require_timetable_manager)
):
    course = db.query(InstitutionalCourse).filter(InstitutionalCourse.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Institutional course not found")
    db.delete(course)
    db.commit()
    return {"status": "deleted"}


# ---------------------------------------------------------------------------
# Shared Course CRUD
# ---------------------------------------------------------------------------

@router.post("/shared-courses", response_model=SharedCourseOut, status_code=status.HTTP_201_CREATED)
def create_shared_course(
    payload: SharedCourseCreate,
    db: Session = Depends(get_db),
    _: User = Depends(require_timetable_manager)
):
    if db.query(User).filter(User.id == payload.faculty_id).first() is None:
        raise HTTPException(status_code=404, detail=f"No faculty found with id {payload.faculty_id}")
    if payload.room_id and db.query(Room).filter(Room.id == payload.room_id).first() is None:
        raise HTTPException(status_code=404, detail=f"No room found with id {payload.room_id}")

    course = SharedCourse(**payload.model_dump())
    db.add(course)
    db.commit()
    db.refresh(course)
    return course


@router.get("/shared-courses", response_model=list[SharedCourseOut])
def list_shared_courses(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return db.query(SharedCourse).all()


@router.delete("/shared-courses/{course_id}")
def delete_shared_course(
    course_id: str,
    db: Session = Depends(get_db),
    _: User = Depends(require_timetable_manager)
):
    course = db.query(SharedCourse).filter(SharedCourse.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Shared course not found")
    db.delete(course)
    db.commit()
    return {"status": "deleted"}


# ---------------------------------------------------------------------------
# Pre-flight validation & Solve endpoint
# ---------------------------------------------------------------------------

def _build_generation_request(db: Session) -> TimetableGenerationRequest:
    # 1. Fetch config or default
    config = db.query(ScheduleConfig).filter(ScheduleConfig.id == "default").first()
    if not config:
        config = ScheduleConfig(
            working_days=6,
            periods_per_day=8,
            break_slots=[],
            max_lectures_per_day_per_faculty=None,
            time_limit_seconds=30
        )

    divisions = db.query(Division).all()
    subjects = db.query(Subject).all()
    rooms = db.query(Room).all()
    assignments = db.query(TeachingAssignment).all()

    # Gather legacy unavailability + generic constraints
    unavail_list = []
    legacy_unavail = db.query(FacultyUnavailability).all()
    for u in legacy_unavail:
        unavail_list.append(SolverUnavailability(faculty_id=u.faculty_id, day=u.day, slot=u.slot))

    generic_constraints = db.query(TimetableConstraint).filter(TimetableConstraint.is_active == True).all()
    soft_constraints = []

    for c in generic_constraints:
        payload = c.payload
        if c.priority == "hard" and c.constraint_type == "faculty_unavailability":
            unavail_list.append(SolverUnavailability(
                faculty_id=payload.get("faculty_id"),
                day=int(payload.get("day", 0)),
                slot=int(payload.get("slot", 0))
            ))
        elif c.priority == "soft":
            soft_constraints.append(SolverSoftConstraint(
                type=c.constraint_type,
                payload=payload
            ))

    # Gather institutional fixed courses and shared courses
    institutional_db = db.query(InstitutionalCourse).all()
    shared_db = db.query(SharedCourse).all()

    solver_institutional = [
        SolverInstitutionalCourse(
            course_name=ic.course_name,
            course_code=ic.course_code,
            year=ic.year,
            divisions=ic.divisions,
            day=ic.day,
            start_slot=ic.start_slot,
            duration_slots=ic.duration_slots,
            faculty_id=ic.faculty_id,
            room_id=ic.room_id
        ) for ic in institutional_db
    ]

    solver_shared = [
        SolverSharedCourse(
            id=sc.id,
            course_name=sc.course_name,
            course_code=sc.course_code,
            year=sc.year,
            divisions=sc.divisions,
            faculty_id=sc.faculty_id,
            room_id=sc.room_id,
            duration_slots=sc.duration_slots,
            weekly_sessions=sc.weekly_sessions,
            session_type=sc.session_type
        ) for sc in shared_db
    ]

    return TimetableGenerationRequest(
        divisions=[SolverDivision(id=d.id, name=d.name, division_code=d.division_code, year=d.year, strength=d.strength) for d in divisions],
        subjects=[
            SolverSubject(
                id=s.id, name=s.name, weekly_lectures=s.weekly_lectures,
                is_lab=s.is_lab, lab_sessions_per_week=s.lab_sessions_per_week,
                lab_block_size=s.lab_block_size
            ) for s in subjects
        ],
        rooms=[SolverRoom(id=r.id, name=r.name, type=r.type, capacity=r.capacity) for r in rooms],
        assignments=[
            SolverAssignment(faculty_id=a.faculty_id, subject_id=a.subject_id, division_id=a.division_id)
            for a in assignments
        ],
        unavailability=unavail_list,
        institutional_courses=solver_institutional,
        shared_courses=solver_shared,
        working_days=config.working_days or 6,
        periods_per_day=config.periods_per_day or 8,
        break_slots=config.break_slots or [],
        max_lectures_per_day_per_faculty=config.max_lectures_per_day_per_faculty,
        time_limit_seconds=config.time_limit_seconds or 30,
        soft_constraints=soft_constraints
    )



@router.get("/validate", response_model=ValidationResponse)
def pre_validate(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    """Performs pre-solve conflict check against the current DB data."""
    request = _build_generation_request(db)
    conflicts = validate_request(request)
    
    valid = len(conflicts) == 0
    summary = "Configuration is valid for solver run." if valid else f"Found {len(conflicts)} scheduling conflicts."
    suggestions = []
    for c in conflicts:
        if c.type == "no_compatible_lecture_room":
            suggestions.append(f"Add a lecture room matching subject capacity needs.")
        elif c.type == "no_compatible_lab_room":
            suggestions.append(f"Add a lab-type room for {c.subject}.")
        elif c.type == "faculty_overloaded":
            suggestions.append(f"Decrease session requirement or add another faculty for Faculty {c.faculty}.")

    return ValidationResponse(
        valid=valid,
        conflicts=conflicts,
        suggestions=list(set(suggestions)),
        summary=summary
    )


@router.post("/generate")
def generate(
    db: Session = Depends(get_db),
    _: User = Depends(require_timetable_manager),
):
    """
    Loads all current divisions/subjects/rooms/assignments/constraints,
    runs the CP-SAT solver, performs post-generation validation,
    and replaces the existing schedule entries.
    Records history in GenerationRun.
    """
    # 1. Build solver request
    request = _build_generation_request(db)
    
    # Snapshot of config
    config_dict = {
        "working_days": request.working_days,
        "periods_per_day": request.periods_per_day,
        "break_slots": request.break_slots,
        "max_lectures_per_day_per_faculty": request.max_lectures_per_day_per_faculty
    }

    # 2. Run solver
    result = generate_timetable(request)

    batch_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)

    # 3. Handle solver result
    if result.status not in ("OPTIMAL", "FEASIBLE"):
        # Save run history as INFEASIBLE
        db_run = GenerationRun(
            id=batch_id,
            status=result.status,
            solve_time_seconds=result.solve_time_seconds,
            total_entries=0,
            validation_passed=False,
            config_snapshot=config_dict,
            conflicts=[c.model_dump() for c in result.conflicts],
            suggestions=result.suggestions,
            solver_log=result.solver_log
        )
        db.add(db_run)
        db.commit()

        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail={
                "status": result.status,
                "message": result.message or "Solver failed to find a valid solution.",
                "conflicts": [c.model_dump() for c in result.conflicts],
                "suggestions": result.suggestions
            }
        )

    # 4. Run post-solve validation layer
    validation_passed, validation_violations = validate_generated_timetable(request, result.entries)

    if not validation_passed:
        # Validation layer failed (strict audit block)
        db_run = GenerationRun(
            id=batch_id,
            status="ERROR",
            solve_time_seconds=result.solve_time_seconds,
            total_entries=0,
            validation_passed=False,
            config_snapshot=config_dict,
            conflicts=[c.model_dump() for c in validation_violations],
            suggestions=["Check model solver constraints integrity."],
            solver_log=result.solver_log
        )
        db.add(db_run)
        db.commit()

        raise HTTPException(
            status_code=500,
            detail={
                "status": "VALIDATION_FAILED",
                "message": "Independent validation of generated timetable failed.",
                "conflicts": [c.model_dump() for c in validation_violations]
            }
        )

    # 5. Clear old timetable entries and insert new ones
    db.query(TimetableEntry).delete()

    for entry in result.entries:
        db.add(TimetableEntry(
            batch_id=batch_id,
            division_id=entry.division_id,
            subject_id=entry.subject_id,
            faculty_id=entry.faculty_id,
            room_id=entry.room_id,
            day=entry.day,
            slot=entry.slot,
            is_lab_block=entry.is_lab_block,
            generated_at=now,
        ))

    # Record successful run history
    db_run = GenerationRun(
        id=batch_id,
        status=result.status,
        solve_time_seconds=result.solve_time_seconds,
        total_entries=len(result.entries),
        objective_score=result.objective_score,
        validation_passed=True,
        config_snapshot=config_dict,
        conflicts=[],
        suggestions=[],
        solver_log=result.solver_log
    )
    db.add(db_run)

    # Notify affected faculty members
    affected_faculty_ids = {entry.faculty_id for entry in result.entries}
    for faculty_id in affected_faculty_ids:
        notify(
            db,
            recipient_id=faculty_id,
            title="Timetable updated",
            message="Your timetable has been regenerated. Open Timetable to see your new schedule.",
        )

    db.commit()

    return {
        "batch_id": batch_id,
        "status": result.status,
        "total_entries": len(result.entries),
        "solve_time_seconds": result.solve_time_seconds,
        "objective_score": result.objective_score,
        "validation_passed": True
    }


# ---------------------------------------------------------------------------
# Generation History / Logs API
# ---------------------------------------------------------------------------

@router.get("/history", response_model=list[GenerationRunOut])
def get_generation_history(db: Session = Depends(get_db), _: User = Depends(require_timetable_manager)):
    """Retrieves all past generation runs."""
    return db.query(GenerationRun).order_by(GenerationRun.generated_at.desc()).all()


# ---------------------------------------------------------------------------
# Seeding / Upload / Deletion APIs
# ---------------------------------------------------------------------------

@router.post("/seed-sample-data")
def seed_sample_data(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_timetable_manager),
):
    unique = uuid.uuid4().hex[:6]

    # Clear config and seed standard one
    db.query(ScheduleConfig).delete()
    config = ScheduleConfig(
        id="default",
        working_days=5,
        day_names=["Tue", "Wed", "Thur", "Fri", "Sat"], # Match Flutter Day Scheme
        periods_per_day=8,
        period_duration_minutes=60,
        start_time="09:00",
        break_slots=[2, 5],
        break_labels={"2": "Short Break", "5": "Lunch Break"},
        college_name="ENOSIS Engineering Institute"
    )
    db.add(config)

    division_a = Division(name=f"Sample-A-{unique}", year=2, division_code="A", strength=60)
    division_b = Division(name=f"Sample-B-{unique}", year=2, division_code="B", strength=55)
    db.add_all([division_a, division_b])

    daa = Subject(name=f"DAA ({unique})", code="CS301", weekly_lectures=3)
    os_subject = Subject(name=f"Operating Systems ({unique})", code="CS302", weekly_lectures=3)
    dbms = Subject(
        name=f"DBMS ({unique})", code="CS303",
        weekly_lectures=2, is_lab=True, lab_sessions_per_week=1, lab_block_size=2,
    )
    maths = Subject(name=f"Engineering Maths ({unique})", code="MA301", weekly_lectures=4)
    db.add_all([daa, os_subject, dbms, maths])

    room_301 = Room(name=f"Room 301-{unique}", type="lecture", capacity=70)
    room_302 = Room(name=f"Room 302-{unique}", type="lecture", capacity=70)
    lab_1 = Room(name=f"Lab 1-{unique}", type="lab", capacity=70)
    db.add_all([room_301, room_302, lab_1])

    db.flush()

    for subject in (daa, os_subject, dbms, maths):
        db.add(TeachingAssignment(faculty_id=current_user.id, subject_id=subject.id, division_id=division_a.id))
        db.add(TeachingAssignment(faculty_id=current_user.id, subject_id=subject.id, division_id=division_b.id))

    db.commit()

    return {
        "status": "seeded",
        "divisions_created": 2,
        "subjects_created": 4,
        "rooms_created": 3,
        "assignments_created": 8,
        "message": f"Sample data created (tagged '{unique}') with default 5-day config and breaks.",
    }


@router.post("/clear-all")
def clear_all(
    db: Session = Depends(get_db),
    _: User = Depends(require_timetable_manager),
):
    db.query(TimetableEntry).delete()
    db.query(FacultyUnavailability).delete()
    db.query(TimetableConstraint).delete()
    db.query(GenerationRun).delete()
    db.query(TeachingAssignment).delete()
    db.query(Division).delete()
    db.query(Subject).delete()
    db.query(Room).delete()
    db.query(ScheduleConfig).delete()
    db.commit()
    return {"message": "All setup data, constraints, and timetable entries cleared successfully."}


def _ensure_placeholder_entities(db: Session, course_name: str, course_code: str | None, faculty_id: str | None, room_id: str | None):
    # 1. Subject
    subject = db.query(Subject).filter(Subject.name == course_name).first()
    if not subject:
        subject = Subject(id=str(uuid.uuid4()), name=course_name, code=course_code or "FIXED")
        db.add(subject)
        db.flush()
    
    # 2. Faculty
    if not faculty_id:
        faculty = db.query(User).first()
        if not faculty:
            # Create a default guest placeholder user
            faculty = User(
                id=str(uuid.uuid4()), 
                email="placeholder@enosis.edu", 
                full_name="Staff / Guest", 
                hashed_password="",
                role="faculty"
            )
            db.add(faculty)
            db.flush()
        faculty_id = faculty.id
        
    # 3. Room
    if not room_id:
        room = db.query(Room).filter(Room.name == "TBA Classroom").first()
        if not room:
            room = Room(id=str(uuid.uuid4()), name="TBA Classroom", type="lecture", capacity=100)
            db.add(room)
            db.flush()
        room_id = room.id
        
    return subject.id, faculty_id, room_id


@router.post("/upload-excel")
async def upload_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_timetable_manager),
):
    filename = file.filename or ""
    content = await file.read()

    # Clear previous database state
    db.query(TimetableEntry).delete()
    db.query(FacultyUnavailability).delete()
    db.query(TimetableConstraint).delete()
    db.query(TeachingAssignment).delete()
    db.query(InstitutionalCourse).delete()
    db.query(SharedCourse).delete()
    db.query(Division).delete()
    db.query(Subject).delete()
    db.query(Room).delete()
    db.query(ScheduleConfig).delete()
    db.flush()

    # Seed Default config alongside Excel upload
    config = ScheduleConfig(
        id="default",
        working_days=6,
        day_names=["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"],
        periods_per_day=8,
        period_duration_minutes=60,
        lecture_duration_minutes=60,
        lab_duration_minutes=120,
        tutorial_duration_minutes=60,
        start_time="09:00"
    )
    db.add(config)

    created_divisions = []
    created_subjects = []
    created_rooms = []
    created_assignments = []
    created_fixed = []
    created_shared = []

    errors = []

    if filename.lower().endswith(".xlsx"):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")

        # 1. Parse Divisions
        if "Divisions" in wb.sheetnames:
            ws = wb["Divisions"]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1)) if cell.value]
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):
                    continue
                row_dict = dict(zip(headers, row))
                try:
                    div = Division(
                        id=str(uuid.uuid4()),
                        name=str(row_dict.get("name", "")).strip(),
                        year=int(row_dict.get("year", 1)),
                        division_code=str(row_dict.get("division_code", "A")).strip(),
                        strength=int(row_dict.get("strength", 60)),
                    )
                    db.add(div)
                    created_divisions.append(div)
                except Exception as e:
                    errors.append(f"Divisions Sheet Row {row_idx}: {str(e)}")

        # 2. Parse Subjects
        if "Subjects" in wb.sheetnames:
            ws = wb["Subjects"]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1)) if cell.value]
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):
                    continue
                row_dict = dict(zip(headers, row))
                try:
                    sub = Subject(
                        id=str(uuid.uuid4()),
                        name=str(row_dict.get("name", "")).strip(),
                        code=str(row_dict.get("code", "")).strip(),
                        weekly_lectures=int(row_dict.get("weekly_lectures", 3)),
                        is_lab=bool(row_dict.get("is_lab", False)),
                        lab_sessions_per_week=int(row_dict.get("lab_sessions_per_week", 0)),
                        lab_block_size=int(row_dict.get("lab_block_size", 2)),
                    )
                    db.add(sub)
                    created_subjects.append(sub)
                except Exception as e:
                    errors.append(f"Subjects Sheet Row {row_idx}: {str(e)}")

        # 3. Parse Rooms
        if "Rooms" in wb.sheetnames:
            ws = wb["Rooms"]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1)) if cell.value]
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):
                    continue
                row_dict = dict(zip(headers, row))
                try:
                    room = Room(
                        id=str(uuid.uuid4()),
                        name=str(row_dict.get("name", "")).strip(),
                        type=str(row_dict.get("type", "lecture")).strip().lower(),
                        capacity=int(row_dict.get("capacity", 60)),
                    )
                    db.add(room)
                    created_rooms.append(room)
                except Exception as e:
                    errors.append(f"Rooms Sheet Row {row_idx}: {str(e)}")

        # Flush to generate IDs
        db.flush()

        # Maps for quick lookups
        div_by_name = {d.name.upper(): d.id for d in created_divisions}
        sub_by_code = {s.code.upper(): s.id for s in created_subjects}
        room_by_name = {r.name.upper(): r.id for r in created_rooms}
        faculty_by_email = {u.email.lower(): u.id for u in db.query(User).all()}

        # 4. Parse Assignments
        if "Assignments" in wb.sheetnames:
            ws = wb["Assignments"]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1)) if cell.value]
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):
                    continue
                row_dict = dict(zip(headers, row))
                try:
                    fac_email = str(row_dict.get("faculty_email", "")).strip().lower()
                    sub_code = str(row_dict.get("subject_code", "")).strip().upper()
                    div_name = str(row_dict.get("division_name", "")).strip().upper()

                    fac_id = faculty_by_email.get(fac_email)
                    sub_id = sub_by_code.get(sub_code)
                    div_id = div_by_name.get(div_name)

                    if not fac_id:
                        errors.append(f"Assignments Row {row_idx}: Faculty email '{fac_email}' not found.")
                        continue
                    if not sub_id:
                        errors.append(f"Assignments Row {row_idx}: Subject code '{sub_code}' not found.")
                        continue
                    if not div_id:
                        errors.append(f"Assignments Row {row_idx}: Division name '{div_name}' not found.")
                        continue

                    assignment = TeachingAssignment(
                        id=str(uuid.uuid4()),
                        faculty_id=fac_id,
                        subject_id=sub_id,
                        division_id=div_id,
                    )
                    db.add(assignment)
                    created_assignments.append(assignment)
                except Exception as e:
                    errors.append(f"Assignments Row {row_idx}: {str(e)}")

        # 5. Parse Fixed Courses
        if "FixedCourses" in wb.sheetnames:
            ws = wb["FixedCourses"]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1)) if cell.value]
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):
                    continue
                row_dict = dict(zip(headers, row))
                try:
                    divisions_str = str(row_dict.get("divisions", ""))
                    div_list = [d.strip() for d in divisions_str.split(",") if d.strip()]

                    fac_email = str(row_dict.get("faculty_email", "")).strip().lower() if row_dict.get("faculty_email") else None
                    rm_name = str(row_dict.get("room_name", "")).strip().upper() if row_dict.get("room_name") else None

                    fac_id = faculty_by_email.get(fac_email) if fac_email else None
                    rm_id = room_by_name.get(rm_name) if rm_name else None

                    if fac_email and not fac_id:
                        errors.append(f"FixedCourses Row {row_idx}: Faculty email '{fac_email}' not found.")
                        continue
                    if rm_name and not rm_id:
                        errors.append(f"FixedCourses Row {row_idx}: Room '{rm_name}' not found.")
                        continue

                    course = InstitutionalCourse(
                        id=str(uuid.uuid4()),
                        course_name=str(row_dict.get("course_name", "")).strip(),
                        course_code=str(row_dict.get("course_code", "")) or None,
                        year=int(row_dict.get("year", 1)),
                        divisions=div_list,
                        day=int(row_dict.get("day", 0)),
                        start_slot=int(row_dict.get("start_slot", 0)),
                        duration_slots=int(row_dict.get("duration_slots", 1)),
                        faculty_id=fac_id,
                        room_id=rm_id
                    )
                    
                    # Ensure placeholder entities if they are Null so timetable foreign keys pass
                    resolved_sub, resolved_fac, resolved_room = _ensure_placeholder_entities(
                        db, course.course_name, course.course_code, course.faculty_id, course.room_id
                    )
                    course.faculty_id = resolved_fac
                    course.room_id = resolved_room

                    db.add(course)
                    created_fixed.append(course)
                except Exception as e:
                    errors.append(f"FixedCourses Row {row_idx}: {str(e)}")

        # 6. Parse Shared Courses
        if "SharedCourses" in wb.sheetnames:
            ws = wb["SharedCourses"]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1)) if cell.value]
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):
                    continue
                row_dict = dict(zip(headers, row))
                try:
                    divisions_str = str(row_dict.get("divisions", ""))
                    div_list = [d.strip() for d in divisions_str.split(",") if d.strip()]

                    fac_email = str(row_dict.get("faculty_email", "")).strip().lower()
                    rm_name = str(row_dict.get("room_name", "")).strip().upper() if row_dict.get("room_name") else None

                    fac_id = faculty_by_email.get(fac_email)
                    rm_id = room_by_name.get(rm_name) if rm_name else None

                    if not fac_id:
                        errors.append(f"SharedCourses Row {row_idx}: Faculty email '{fac_email}' not found.")
                        continue
                    if rm_name and not rm_id:
                        errors.append(f"SharedCourses Row {row_idx}: Room '{rm_name}' not found.")
                        continue

                    course = SharedCourse(
                        id=str(uuid.uuid4()),
                        course_name=str(row_dict.get("course_name", "")).strip(),
                        course_code=str(row_dict.get("course_code", "")) or None,
                        year=int(row_dict.get("year", 1)),
                        divisions=div_list,
                        faculty_id=fac_id,
                        room_id=rm_id,
                        duration_slots=int(row_dict.get("duration_slots", 1)),
                        weekly_sessions=int(row_dict.get("weekly_sessions", 1)),
                        session_type=str(row_dict.get("session_type", "lecture")).strip().lower()
                    )
                    
                    # Ensure placeholder subject exists so we can map it
                    resolved_sub, _, _ = _ensure_placeholder_entities(
                        db, course.course_name, course.course_code, course.faculty_id, course.room_id
                    )

                    db.add(course)
                    created_shared.append(course)
                except Exception as e:
                    errors.append(f"SharedCourses Row {row_idx}: {str(e)}")

        if errors:
            db.rollback()
            raise HTTPException(status_code=400, detail={"message": "Validation errors found in Excel sheets.", "errors": errors})

    elif filename.lower().endswith(".csv"):
        # Legacy CSV parser for backwards compatibility
        text = content.decode("utf-8")
        reader = csv.DictReader(io.StringIO(text))
        for row_dict in reader:
            if "name" in row_dict:
                sub = Subject(
                    id=str(uuid.uuid4()),
                    name=row_dict["name"].strip(),
                    code=row_dict.get("code", "").strip(),
                    weekly_lectures=int(row_dict.get("weekly_lectures", 3)),
                    is_lab=row_dict.get("is_lab", "").lower() in ("true", "1", "yes"),
                    lab_sessions_per_week=int(row_dict.get("lab_sessions_per_week", 0)),
                    lab_block_size=int(row_dict.get("lab_block_size", 1)),
                )
                db.add(sub)
                created_subjects.append(sub)
    else:
        raise HTTPException(status_code=400, detail="Only .xlsx and .csv files are supported.")

    db.commit()

    return {
        "message": "Excel configuration uploaded and parsed successfully.",
        "divisions_created": len(created_divisions),
        "subjects_created": len(created_subjects),
        "rooms_created": len(created_rooms),
        "assignments_created": len(created_assignments),
        "fixed_courses_created": len(created_fixed),
        "shared_courses_created": len(created_shared)
    }


# ---------------------------------------------------------------------------
# Read-only viewing — this is what Android actually calls.
# ---------------------------------------------------------------------------

def _to_entry_out(entry: TimetableEntry) -> TimetableEntryOut:
    return TimetableEntryOut(
        id=entry.id,
        batch_id=entry.batch_id,
        day=entry.day,
        slot=entry.slot,
        is_lab_block=entry.is_lab_block,
        division_name=entry.division.name,
        division_year=entry.division.year,
        division_code=entry.division.division_code,
        subject_name=entry.subject.name,
        faculty_name=entry.faculty.full_name,
        room_name=entry.room.name,
        generated_at=entry.generated_at,
    )


@router.get("/me", response_model=list[TimetableEntryOut])
def my_timetable(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    entries = db.query(TimetableEntry).filter(TimetableEntry.faculty_id == current_user.id).all()
    return [_to_entry_out(e) for e in entries]


@router.get("/division/{division_id}", response_model=list[TimetableEntryOut])
def division_timetable(division_id: str, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    if division_id == "all":
        entries = db.query(TimetableEntry).all()
        return [_to_entry_out(e) for e in entries]

    division = db.query(Division).filter(Division.id == division_id).first()
    if division is None:
        raise HTTPException(status_code=404, detail="Division not found")

    entries = db.query(TimetableEntry).filter(TimetableEntry.division_id == division_id).all()
    return [_to_entry_out(e) for e in entries]


@router.get("/year/{year}", response_model=list[TimetableEntryOut])
def year_timetable(year: int, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    division_ids = [d.id for d in db.query(Division).filter(Division.year == year).all()]
    if not division_ids:
        return []

    entries = db.query(TimetableEntry).filter(TimetableEntry.division_id.in_(division_ids)).all()
    return [_to_entry_out(e) for e in entries]
